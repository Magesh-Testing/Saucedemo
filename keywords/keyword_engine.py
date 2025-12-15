from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from Saucedemo.keywords.login_keywords import LoginKeywords
import logging

logger = logging.getLogger("SauceDemoLogger")

class KeywordEngine:

    def execute(self, excl_path):
        logger.info(f"Executing keywords from file: {excl_path}")

        workbook = load_workbook(excl_path)
        sheet = workbook.active

        driver = None
        lk = None

        for row in sheet.iter_rows(min_row=2, values_only=True):
            #row = list(row) + ["", "","",""]
            testcase, keyword, locator, testdata = row
            logger.info(f"Keyword - {keyword}, Locator - {locator}, TestData - {testdata}")

            #Open browser
            if keyword == "open_browser":
                logger.info(f"Opening browser: {testdata}")
                if testdata == "chrome":
                    driver = webdriver.Chrome()
                else:
                    logger.warning(f"Unknown browser '{testdata}', defaulting to firefox.")
                    driver = webdriver.Firefox()
                driver.maximize_window()
                lk = LoginKeywords(driver)

            #Open URL
            elif keyword == "open_url":
                logger.info(f"Navigating to URL: {testdata}")
                driver.get(testdata)

            elif keyword == "input_text":
                if locator:
                    lk.input_text(locator, testdata)
                else:
                    logger.error("Locator missing for input_text")

            elif keyword == "click":
                if locator:
                    lk.click(locator)
                else:
                    logger.error("Locator missing for click functionality")
            elif keyword == "verify_text":
                if locator:
                    lk.verify_text(locator, testdata)
                else:
                    logger.error("Locator missing for verify_text")
            elif keyword == "close_browser":
                logger.info("Closing browser.")
                driver.quit()
        return driver