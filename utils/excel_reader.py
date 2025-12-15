from openpyxl import load_workbook

def read_excel(filepath, filter_status=None):
    workbook = load_workbook(filepath)
    sheet = workbook.active

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        username, password, expected, status  = row

        # If filter is applied -> return only matching rows
        if filter_status:
            if status and status.lower() == filter_status.lower():
                data.append((username, password, expected, status))
        else:
            data.append((username, password, expected, status))

        #data.append((username, password, expected, status))
    return data